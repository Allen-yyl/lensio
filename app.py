import streamlit as st
import cv2
import os
import tempfile
import json
import base64
import zipfile
import io
import shutil
from openai import OpenAI
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Lensio短视频分镜拉片工具", layout="wide")

@st.cache_resource
def get_client():
    return OpenAI(
        api_key="sk-73affc2fea0b46dc8062facde27bbdc3",
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
    )

client = get_client()
VISION_MODEL = "qwen3-vl-flash"

def seconds_to_time(s):
    m = int(s // 60)
    sec = int(s % 60)
    return f"{m:02d}:{sec:02d}"

def time_to_seconds(t):
    parts = t.split(':')
    return int(parts[0]) * 60 + int(parts[1])

def find_closest_frame_idx(start_time_str, frame_times):
    """Find the frame index closest to the given start time"""
    target_sec = time_to_seconds(start_time_str)
    closest_idx = 0
    min_diff = abs(frame_times[0] - target_sec)
    for i, ft in enumerate(frame_times):
        diff = abs(ft - target_sec)
        if diff < min_diff:
            min_diff = diff
            closest_idx = i
    return closest_idx

def extract_frames(video_path, output_dir, fps=1):
    cap = cv2.VideoCapture(video_path)
    video_fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    duration = total_frames / video_fps
    frame_paths = []
    frame_times = []
    frame_idx = 0
    saved_idx = 0
    interval = max(1, int(video_fps / fps))
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if frame_idx % interval == 0:
            frame_path = os.path.join(output_dir, f"frame_{saved_idx:04d}.jpg")
            encode_param = [int(cv2.IMWRITE_JPEG_QUALITY), 100]
            cv2.imwrite(frame_path, frame, encode_param)
            frame_paths.append(frame_path)
            frame_times.append(saved_idx / fps)
            saved_idx += 1
        frame_idx += 1
    cap.release()
    return frame_paths, frame_times, duration

def load_frames_as_base64(frame_paths):
    frames_b64 = []
    for path in frame_paths:
        with open(path, 'rb') as f:
            data = f.read()
        frames_b64.append(base64.b64encode(data).decode('utf-8'))
    return frames_b64

def create_shots_zip(frames_b64, frame_times, results):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for idx, result in enumerate(results, 1):
            start_str = result['start'].replace(':', '')
            end_str = result['end'].replace(':', '')
            frame_idx = find_closest_frame_idx(result['start'], frame_times)
            if 0 <= frame_idx < len(frames_b64):
                shot_name = f"shot_{idx:02d}_{start_str}-{end_str}_{result.get('shot_type', '')}.jpg"
                zf.writestr(shot_name, base64.b64decode(frames_b64[frame_idx]))
    zip_buffer.seek(0)
    return zip_buffer

def create_shots_excel(frames_b64, frame_times, results):
    """Create an Excel file with shot data and embedded screenshots"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分镜分析"

    # Headers
    headers = ["序号", "截图", "开始", "结束", "场景环境", "主体", "景别", "机位角度", "分镜内容"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    col_widths = {1: 6, 2: 15, 3: 8, 4: 8, 5: 25, 6: 20, 7: 8, 8: 15, 9: 40}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Data rows
    img_temp_dir = tempfile.mkdtemp()
    try:
        for idx, result in enumerate(results, 1):
            row = idx + 1

            # Write text data
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=3, value=result['start'])
            ws.cell(row=row, column=4, value=result['end'])
            ws.cell(row=row, column=5, value=result.get('scene', ''))
            ws.cell(row=row, column=6, value=result.get('subject', ''))
            ws.cell(row=row, column=7, value=result.get('shot_type', ''))
            ws.cell(row=row, column=8, value=result.get('camera_angle', ''))
            ws.cell(row=row, column=9, value=result.get('description', ''))

            # Embed image
            frame_idx = find_closest_frame_idx(result['start'], frame_times)
            if 0 <= frame_idx < len(frames_b64):
                img_path = os.path.join(img_temp_dir, f"shot_{idx}.jpg")
                with open(img_path, 'wb') as f:
                    f.write(base64.b64decode(frames_b64[frame_idx]))

                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(img_path)
                img.width = 120
                img.height = 68
                ws.add_image(img, f'B{row}')

            # Apply alignment
            for col in range(1, 10):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

            ws.row_dimensions[row].height = 60

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer
    finally:
        shutil.rmtree(img_temp_dir, ignore_errors=True)

def extract_location_keyword(scene_desc):
    """Extract the most specific spatial location keyword"""
    location_keywords = ['过道', '走廊', '门框', '门口', '窗边', '车窗', '座位', '车窗边']
    for kw in location_keywords:
        if kw in scene_desc:
            return kw
    return scene_desc[:10] if scene_desc else ""

def render_shot_row(idx, frames_b64, frame_times, result):
    frame_idx = find_closest_frame_idx(result['start'], frame_times)

    cols = st.columns([0.04, 0.16, 0.11, 0.18, 0.16, 0.10, 0.10, 0.15])
    with cols[0]:
        st.markdown(f"**{idx}**")
    with cols[1]:
        if 0 <= frame_idx < len(frames_b64):
            st.image(f"data:image/jpeg;base64,{frames_b64[frame_idx]}", width=110)
    with cols[2]:
        st.markdown(f"`{result['start']}`<br>`{result['end']}`", unsafe_allow_html=True)
    with cols[3]:
        st.markdown(result.get('scene', ''))
    with cols[4]:
        st.markdown(result.get('subject', ''))
    with cols[5]:
        st.markdown(result.get('shot_type', ''))
    with cols[6]:
        st.markdown(result.get('camera_angle', ''))
    with cols[7]:
        st.markdown(result.get('description', ''))

    st.divider()

def render_results(shots, frames_b64, frame_times, video_summary):
    if not shots:
        return

    st.subheader(f"分镜分析结果 (共 {len(shots)} 个分镜)")

    header_cols = st.columns([0.04, 0.16, 0.11, 0.18, 0.16, 0.10, 0.10, 0.15])
    with header_cols[0]: st.markdown("**序号**")
    with header_cols[1]: st.markdown("**截图**")
    with header_cols[2]: st.markdown("**时间范围**")
    with header_cols[3]: st.markdown("**场景环境**")
    with header_cols[4]: st.markdown("**主体**")
    with header_cols[5]: st.markdown("**景别**")
    with header_cols[6]: st.markdown("**机位角度**")
    with header_cols[7]: st.markdown("**分镜内容**")
    st.divider()

    for idx, shot in enumerate(shots, 1):
        render_shot_row(idx, frames_b64, frame_times, shot)

    if video_summary:
        st.divider()
        st.subheader("视频整体总结")
        st.info(video_summary)

COMPARISON_TEMPLATE = """比较当前帧与参考分镜的特征。

参考分镜：
- 场景空间位置: {prev_scene}
- 主体空间姿态: {prev_pose}
- 机位角度: {prev_camera}

当前帧时间: {time_str}

请直接描述当前帧的这三个维度：
返回JSON：{{"scene_now": "场景位置（具体在哪里）", "pose_now": "姿态（坐着/站立/行走等）", "camera_now": "机位角度", "shot_type_now": "景别（全景/中景/近景/特写）"}}"""

SHOT_TEMPLATE_WITH_CONTEXT = """描述这个新分镜的画面内容。

前几个分镜参考：
{prev_shots}

当前帧时间: {time_str}

请描述当前分镜的画面内容，承接前面的剧情发展：
返回JSON：{{"scene": "场景空间位置", "subject_pose": "主体空间姿态", "camera_angle": "机位角度", "shot_type": "景别（全景/中景/近景/特写）", "description": "分镜内容描述100字以内"}}"""

def analyze_frame_with_shot_logic(frames_b64, frame_times, prev_shot):
    """Analyze frames one by one using location/pose/camera based logic"""
    shots = []
    total = len(frames_b64)

    for i in range(total):
        img_b64 = frames_b64[i]
        time_str = seconds_to_time(frame_times[i])

        if i == 0:
            # Ask model to describe first frame
            response = client.chat.completions.create(
                model=VISION_MODEL,
                messages=[{"role": "user", "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                    {"type": "text", "text": """分析这张截图。

返回JSON（不要包含其他文字）：
{"scene": "场景空间位置（具体在哪里，如：窗边/过道/门框/走廊等）", "subject_pose": "主体空间姿态（坐着/站立/行走/躺卧）", "camera_angle": "机位角度（正面/侧面/背面/低角度仰拍等）", "shot_type": "景别（全景/中景/近景/特写）", "description": "分镜内容描述100字以内"}"""}
                ]}],
                max_tokens=1500
            )
            text = response.choices[0].message.content.strip()
            start = text.find('{')
            end = text.rfind('}') + 1
            if start != -1 and end != 0:
                parsed = json.loads(text[start:end])
                shots.append({
                    "start": time_str, "end": time_str,
                    "scene": parsed.get("scene", ""),
                    "subject": parsed.get("subject_pose", ""),
                    "shot_type": parsed.get("shot_type", ""),
                    "camera_angle": parsed.get("camera_angle", ""),
                    "description": parsed.get("description", "")
                })
            yield shots.copy(), i, total
        else:
            prev = shots[-1]

            # Ask model to describe current frame relative to previous
            response = client.chat.completions.create(
                model=VISION_MODEL,
                messages=[{"role": "user", "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                    {"type": "text", "text": COMPARISON_TEMPLATE.format(
                        prev_scene=prev['scene'],
                        prev_pose=prev['subject'],
                        prev_camera=prev['camera_angle'],
                        time_str=time_str
                    )}
                ]}],
                max_tokens=1500
            )

            text = response.choices[0].message.content.strip()
            start = text.find('{')
            end = text.rfind('}') + 1
            if start != -1 and end != 0:
                parsed_curr = json.loads(text[start:end])

                curr_scene = parsed_curr.get('scene_now', '')
                curr_pose = parsed_curr.get('pose_now', '')
                curr_camera = parsed_curr.get('camera_now', '')

                # Extract location keywords
                prev_loc = extract_location_keyword(prev['scene'])
                curr_loc = extract_location_keyword(curr_scene)

                # Pose transitions
                pose_kw_map = {'坐': '坐', '站': '站', '走': '走', '行': '行', '立': '站'}
                prev_pose_kw = [pose_kw_map.get(c) for c in prev['subject'] if c in pose_kw_map]
                curr_pose_kw = [pose_kw_map.get(c) for c in curr_pose if c in pose_kw_map]
                prev_pose_main = prev_pose_kw[0] if prev_pose_kw else None
                curr_pose_main = curr_pose_kw[0] if curr_pose_kw else None
                is_pose_transition = bool(prev_pose_main and curr_pose_main and prev_pose_main != curr_pose_main)

                # Camera transitions
                is_camera_transition = False
                prev_is_looking_up = any(k in prev['camera_angle'] for k in ['仰拍', '俯拍', '仰视'])
                curr_is_looking_up = any(k in curr_camera for k in ['仰拍', '俯拍', '仰视'])
                if curr_is_looking_up != prev_is_looking_up:
                    is_camera_transition = True

                prev_is_back = any(k in prev['camera_angle'] for k in ['背面', '背对'])
                curr_is_back = any(k in curr_camera for k in ['背面', '背对'])
                if curr_is_back != prev_is_back:
                    is_camera_transition = True

                # Location transitions
                is_location_transition = (curr_loc != prev_loc and len(curr_loc) > 1 and len(prev_loc) > 1)

                is_new = is_location_transition or is_pose_transition or is_camera_transition

                if is_new:
                    # Build context from previous shots
                    prev_shots_text = "\n".join([
                        f"分镜{idx}: {s['scene']} | {s['subject']} | {s['description']}"
                        for idx, s in enumerate(shots[-3:], len(shots)-2)
                    ]) if len(shots) > 0 else "（首个分镜）"

                    response2 = client.chat.completions.create(
                        model=VISION_MODEL,
                        messages=[{"role": "user", "content": [
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                            {"type": "text", "text": SHOT_TEMPLATE_WITH_CONTEXT.format(
                                prev_shots=prev_shots_text,
                                time_str=time_str
                            )}
                        ]}],
                        max_tokens=1500
                    )
                    text2 = response2.choices[0].message.content.strip()
                    start2 = text2.find('{')
                    end2 = text2.rfind('}') + 1
                    if start2 != -1 and end2 != 0:
                        parsed_new = json.loads(text2[start2:end2])
                        shots.append({
                            "start": time_str, "end": time_str,
                            "scene": parsed_new.get("scene", curr_scene),
                            "subject": parsed_new.get("subject_pose", curr_pose),
                            "shot_type": parsed_new.get("shot_type", ""),
                            "camera_angle": parsed_new.get("camera_angle", curr_camera),
                            "description": parsed_new.get("description", "")
                        })
                else:
                    shots[-1]["end"] = time_str

            yield shots.copy(), i, total

def main():
    st.markdown("""
    <style>
    .main-title { font-size: 1.8rem !important; }
    .sub-title { font-size: 1rem !important; color: #666; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div style="position: fixed; top: 10px; left: 20px; font-size: 0.85rem; color: #888; z-index: 9999; background: rgba(255,255,255,0.8); padding: 2px 6px; border-radius: 4px;">created by 力包 For 莉宝</div>', unsafe_allow_html=True)
    st.markdown('<p class="main-title">Lensio</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-title">短视频分镜拉片工具</p>', unsafe_allow_html=True)

    defaults = {
        'shots': [], 'frames_b64': [], 'frame_times': [],
        'analyzing': False, 'video_summary': '',
        'phase': 'idle',
        'current_frame_idx': 0, 'total_frames': 0,
        'temp_dir': None
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    if st.session_state.phase == 'extracting':
        st.progress(5, text="正在截取视频帧...")
    elif st.session_state.phase == 'analyzing':
        pct = int(st.session_state.current_frame_idx / st.session_state.total_frames * 90) if st.session_state.total_frames > 0 else 0
        st.progress(pct, text=f"正在分析中 ({pct}%)")
    elif st.session_state.phase == 'summarizing':
        st.progress(95, text="正在生成视频整体总结...")
    elif st.session_state.phase == 'done':
        st.progress(100, text="分析完成")

    uploaded_file = st.file_uploader(
        "拖拽视频文件到此处",
        type=['mov', 'mp4', 'avi', 'mkv'],
        disabled=st.session_state.phase != 'idle',
        label_visibility="collapsed"
    )

    button_disabled = uploaded_file is None or st.session_state.phase != 'idle'
    start_clicked = st.button("开始分析", type="primary", disabled=button_disabled)

    if start_clicked and uploaded_file:
        st.session_state.shots = []
        st.session_state.video_summary = ""
        st.session_state.phase = 'extracting'

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.mov')
        temp_file.write(uploaded_file.getvalue())
        temp_file.close()

        st.session_state.temp_dir = tempfile.mkdtemp()
        frame_paths, frame_times, duration = extract_frames(temp_file.name, st.session_state.temp_dir, fps=1)
        st.session_state.frames_b64 = load_frames_as_base64(frame_paths)
        st.session_state.frame_times = frame_times
        st.session_state.total_frames = len(frame_paths)
        st.session_state.current_frame_idx = 0

        os.unlink(temp_file.name)
        shutil.rmtree(st.session_state.temp_dir, ignore_errors=True)
        st.session_state.temp_dir = None

        st.session_state.phase = 'analyzing'
        st.rerun()

    elif st.session_state.phase == 'analyzing':
        i = st.session_state.current_frame_idx
        total = st.session_state.total_frames

        if i < total:
            img_b64 = st.session_state.frames_b64[i]
            time_str = seconds_to_time(st.session_state.frame_times[i])

            if i == 0:
                prompt = """分析这张截图。

返回JSON（不要包含其他文字）：
{"scene": "场景空间位置（具体在哪里，如：窗边/过道/门框/走廊等）", "subject_pose": "主体空间姿态（坐着/站立/行走/躺卧）", "camera_angle": "机位角度（正面/侧面/背面/低角度仰拍等）", "shot_type": "景别（全景/中景/近景/特写）", "description": "分镜内容描述100字以内"}"""
            else:
                prev = st.session_state.shots[-1]
                prompt = COMPARISON_TEMPLATE.format(
                    prev_scene=prev['scene'],
                    prev_pose=prev['subject'],
                    prev_camera=prev['camera_angle'],
                    time_str=time_str
                )

            try:
                response = client.chat.completions.create(
                    model=VISION_MODEL,
                    messages=[{"role": "user", "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                        {"type": "text", "text": prompt}
                    ]}],
                    max_tokens=1500
                )

                text = response.choices[0].message.content.strip()
                start = text.find('{')
                end = text.rfind('}') + 1
                if start != -1 and end != 0:
                    parsed = json.loads(text[start:end])

                    if i == 0:
                        st.session_state.shots.append({
                            "start": time_str, "end": time_str,
                            "scene": parsed.get("scene", ""),
                            "subject": parsed.get("subject_pose", ""),
                            "shot_type": parsed.get("shot_type", ""),
                            "camera_angle": parsed.get("camera_angle", ""),
                            "description": parsed.get("description", "")
                        })
                    else:
                        prev = st.session_state.shots[-1]
                        curr_scene = parsed.get('scene_now', '')
                        curr_pose = parsed.get('pose_now', '')
                        curr_camera = parsed.get('camera_now', '')
                        curr_shot_type = parsed.get('shot_type_now', '')

                        prev_loc = extract_location_keyword(prev['scene'])
                        curr_loc = extract_location_keyword(curr_scene)

                        pose_kw_map = {'坐': '坐', '站': '站', '走': '走', '行': '行', '立': '站'}
                        prev_pose_kw = [pose_kw_map.get(c) for c in prev['subject'] if c in pose_kw_map]
                        curr_pose_kw = [pose_kw_map.get(c) for c in curr_pose if c in pose_kw_map]
                        is_pose_transition = bool(prev_pose_kw and curr_pose_kw and prev_pose_kw[0] != curr_pose_kw[0])

                        prev_is_up = any(k in prev['camera_angle'] for k in ['仰拍', '俯拍', '仰视'])
                        curr_is_up = any(k in curr_camera for k in ['仰拍', '俯拍', '仰视'])
                        is_camera_up = curr_is_up != prev_is_up

                        prev_is_back = any(k in prev['camera_angle'] for k in ['背面', '背对'])
                        curr_is_back = any(k in curr_camera for k in ['背面', '背对'])
                        is_camera_back = curr_is_back != prev_is_back

                        is_location_transition = (curr_loc != prev_loc and len(curr_loc) > 1 and len(prev_loc) > 1)

                        is_new = is_location_transition or is_pose_transition or is_camera_up or is_camera_back

                        if is_new:
                            # Build context from previous shots
                            prev_shots = st.session_state.shots[-3:] if len(st.session_state.shots) >= 3 else st.session_state.shots
                            prev_shots_text = "\n".join([
                                f"分镜{len(st.session_state.shots)-len(prev_shots)+idx+1}: {s['scene']} | {s['subject']} | {s['description']}"
                                for idx, s in enumerate(prev_shots)
                            ]) if prev_shots else "（首个分镜）"

                            response2 = client.chat.completions.create(
                                model=VISION_MODEL,
                                messages=[{"role": "user", "content": [
                                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                                    {"type": "text", "text": SHOT_TEMPLATE_WITH_CONTEXT.format(
                                        prev_shots=prev_shots_text,
                                        time_str=time_str
                                    )}
                                ]}],
                                max_tokens=1500
                            )
                            text2 = response2.choices[0].message.content.strip()
                            start2 = text2.find('{')
                            end2 = text2.rfind('}') + 1
                            if start2 != -1 and end2 != 0:
                                parsed_new = json.loads(text2[start2:end2])
                                st.session_state.shots.append({
                                    "start": time_str, "end": time_str,
                                    "scene": parsed_new.get("scene", curr_scene),
                                    "subject": parsed_new.get("subject_pose", curr_pose),
                                    "shot_type": parsed_new.get("shot_type") or curr_shot_type,
                                    "camera_angle": parsed_new.get("camera_angle", curr_camera),
                                    "description": parsed_new.get("description", "")
                                })
                        else:
                            st.session_state.shots[-1]["end"] = time_str

            except Exception as e:
                if st.session_state.shots:
                    st.session_state.shots[-1]["end"] = time_str

            st.session_state.current_frame_idx += 1
            st.rerun()
        else:
            st.session_state.phase = 'summarizing'
            st.rerun()

    elif st.session_state.phase == 'summarizing':
        if len(st.session_state.shots) > 1:
            shots_summary = "\n".join([
                f"分镜{i+1} ({s['start']}-{s['end']}): 场景={s['scene']}, 主体={s['subject']}, 景别={s['shot_type']}, 机位={s['camera_angle']}, 内容={s['description']}"
                for i, s in enumerate(st.session_state.shots)
            ])

            summary_prompt = f"""根据以下分镜分析，对整个视频进行总结：

{shots_summary}

请用100字左右总结这个视频的整体内容、风格和特点。
请只返回总结文字，不要其他内容。"""

            try:
                response = client.chat.completions.create(
                    model=VISION_MODEL,
                    messages=[{"role": "user", "content": summary_prompt}],
                    max_tokens=300
                )
                st.session_state.video_summary = response.choices[0].message.content.strip()
            except:
                st.session_state.video_summary = f"视频共{len(st.session_state.shots)}个分镜，涵盖多种场景和人物动作。"

        st.session_state.phase = 'done'
        st.rerun()

    if st.session_state.shots:
        render_results(st.session_state.shots, st.session_state.frames_b64, st.session_state.frame_times, st.session_state.video_summary if st.session_state.phase == 'done' else "")

        if st.session_state.phase == 'done':
            st.divider()
            col1, col2 = st.columns([1, 1])

            zip_buffer = create_shots_zip(st.session_state.frames_b64, st.session_state.frame_times, st.session_state.shots)

            with col1:
                st.download_button(
                    "下载所有分镜截图 (ZIP)",
                    zip_buffer,
                    file_name="shots.zip",
                    mime="application/zip",
                    type="secondary"
                )

            with col2:
                excel_buffer = create_shots_excel(st.session_state.frames_b64, st.session_state.frame_times, st.session_state.shots)
                st.download_button(
                    "下载表格 (Excel)",
                    excel_buffer,
                    file_name="shots.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary"
                )

if __name__ == "__main__":
    main()
